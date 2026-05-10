"""
Season 4 QA Harness
===================
Replays the season-4.json dataset through the app's scoring logic and diffs
against Terry's Season 4 spreadsheet (Standings sheet + per-event Results
sheets). Any discrepancy is one of: app bug, Terry math error, or data-entry
error in season-4.json.

Run from the repo root:
    python qa/season4_qa.py

The scoring logic below is a faithful Python port of the JS in
ChubbsMobileApp_v5/index.html — when those functions change, mirror the
change here.
"""
from __future__ import annotations

import json
import math
import os
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

# Force UTF-8 on Windows consoles (default cp1252 chokes on Δ, en-dash, etc.)
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

ROOT = Path(__file__).resolve().parents[1]
SEASON_JSON = ROOT / "ChubbsMobileApp_v5" / "season-4.json"
SHEET_PATH  = ROOT / "Season 4 Scores.xlsx"

DEFAULT_RANK_POINTS = [13, 11, 10, 9, 8, 7, 6, 5, 4, 3, 2, 1]


# ─────────────────────────────────────────────────────────────────────────────
# App scoring logic — Python port of the JS in index.html
# ─────────────────────────────────────────────────────────────────────────────

def shots_received(hcp, si):
    """index.html:3428 — shotsReceived(hcp, si)"""
    safe = max(0, int(hcp or 0))
    base = safe // 18
    rem = safe % 18
    return base + (1 if si <= rem else 0)


def stableford_points(gross, par, shots):
    """index.html:3434 — stablefordPoints(gross, par, shots)
    6/5/4/3/2/1/0 for net diff <=-4 / -3 / -2 / -1 / 0 / +1 / >=+2.
    """
    if gross is None or gross == "":
        return None
    net = gross - shots
    diff = net - par
    if diff <= -4: return 6
    if diff == -3: return 5
    if diff == -2: return 4
    if diff == -1: return 3
    if diff ==  0: return 2
    if diff ==  1: return 1
    return 0


def canon_name(n):
    """index.html:2185 — canonName"""
    if not isinstance(n, str): return n or ""
    return " ".join(n.split())


def aggregate_season(season):
    """index.html:2192 — aggregateSeason(season)
    Returns list of player rows sorted by points desc, with cut-line tiebreak.
    """
    rank_pts = season.get("rankPoints") or DEFAULT_RANK_POINTS
    agg = {}            # name -> dict
    per_player = {}     # name -> list of per-event summaries

    for ev in season["events"]:
        scored = []
        for p in ev["players"]:
            nm = canon_name(p["name"])
            sb = gross_total = dblpar = neagle = nbirdie = 0
            cb = []
            cb_tot = 0
            for i, hole in enumerate(ev["holes"]):
                g = p["gross"][i] if i < len(p["gross"]) else None
                if g is None or g == "": continue
                shots = shots_received(p["hcp"], hole["si"])
                pts = stableford_points(g, hole["par"], shots)
                if pts is not None: sb += pts
                gross_total += g
                diff = (g - shots) - hole["par"]
                if g >= 2 * hole["par"]: dblpar += 1
                if diff <= -2: neagle += 1
                if diff == -1: nbirdie += 1
            for i in range(len(ev["holes"]) - 1, -1, -1):
                g = p["gross"][i] if i < len(p["gross"]) else None
                cb_tot += g if isinstance(g, (int, float)) else 0
                cb.append(cb_tot)
            scored.append({
                "name": nm, "stableford": sb, "gross": gross_total,
                "dblpar": dblpar, "neagle": neagle, "nbirdie": nbirdie, "cb": cb,
            })
        # Stableford desc, then gross-countback ascending (lower is better)
        def _sort_key(r):
            return (-r["stableford"], *r["cb"])
        scored.sort(key=_sort_key)

        for rank, row in enumerate(scored):
            nm = row["name"]
            if nm not in agg:
                agg[nm] = {"name": nm, "played": 0, "stableford": 0,
                           "dblpar": 0, "neagle": 0, "nbirdie": 0}
                per_player[nm] = []
            ep = next((x for x in ev["players"] if canon_name(x["name"]) == nm), None)
            override = ep.get("overrideSeasonPts") if ep else None
            if override is not None and not isinstance(override, bool):
                season_pts = int(override)
            else:
                season_pts = rank_pts[rank] if rank < len(rank_pts) else 0
            agg[nm]["played"]     += 1
            agg[nm]["stableford"] += row["stableford"]
            agg[nm]["dblpar"]     += row["dblpar"]
            agg[nm]["neagle"]     += row["neagle"]
            agg[nm]["nbirdie"]    += row["nbirdie"]
            per_player[nm].append({
                "eventId": ev["id"], "eventName": ev["name"],
                "stableford": row["stableford"], "gross": row["gross"],
                "rank": rank + 1, "seasonPts": season_pts,
            })

    total_events = len(season["events"])
    PLAYOFF_MIN = 3

    rows = []
    for a in agg.values():
        evs = per_player[a["name"]]
        best7 = sum(sorted([e["seasonPts"] for e in evs], reverse=True)[:7])
        perfect10_bonus = 3 if a["played"] == 10 else 0
        points = best7 + perfect10_bonus
        seed_points = sum(e["seasonPts"] for e in evs)
        round_avg = (best7 / a["played"]) if a["played"] else 0
        rows.append({
            **a, "points": points, "seedPoints": seed_points,
            "best7": best7, "perfect10Bonus": perfect10_bonus,
            "roundAvg": round_avg, "events": evs,
            "qualified": a["played"] >= PLAYOFF_MIN,
        })

    # Cut-line tiebreak per Terry 2026-05-10: points desc → round-avg desc →
    # seedPoints desc → stableford desc. Replaces the played-desc rule from
    # 2026-04-29 (Terry didn't remember locking it, confirmed average is what
    # the spreadsheet uses).
    rows.sort(key=lambda r: (-r["points"], -r["roundAvg"], -r["seedPoints"], -r["stableford"]))
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Terry name normalization — apply same rules the Python extractor uses,
# so spreadsheet rows can be compared to canonicalized JSON rows.
# ─────────────────────────────────────────────────────────────────────────────

# Per CLAUDE.md normalization rules
TERRY_TO_CANON = {
    "Mike H":   "Hanson",
    "Mike":     "Mike W",
    "Stuartom": "Stuart",
    "Ryan":     "Ryan N",
    "PLZ":      "Penglei",
    "PengLei":  "Penglei",
    "Greame":   "Graeme",
    # Matts: bare "Matt" → Matt D when Matt D not in same event; OtherMatt/Mathew → Matthew SA
}


def normalize_terry(name, event_player_set=None):
    n = (name or "").strip()
    if n in TERRY_TO_CANON:
        return TERRY_TO_CANON[n]
    # Other Matt / Mathew = Matthew SA (per CLAUDE.md disambiguators).
    if n in ("Other Matt", "Mathew"):
        return "Matthew SA"
    # Bare "Matt": context-aware. If event_player_set already contains "Matt D",
    # then "Matt" must mean the other one (Matthew SA). Otherwise it's Matt D.
    # In sheets with no context (per-event sheets where only one Matt appears),
    # default to Matt D — the regular Season 4 player.
    if n == "Matt":
        if event_player_set is not None and "Matt D" in event_player_set:
            return "Matthew SA"
        return "Matt D"
    return n


# ─────────────────────────────────────────────────────────────────────────────
# Per-event sheet mapping — discovered by matching top-stableford per event
# vs each Sheet* tab in the spreadsheet.
# ─────────────────────────────────────────────────────────────────────────────

SHEET_TO_EVENT = {
    "Sept Results": "sept-birds",
    "Oct Results":  "oct-mountain",
    "Sheet6":       "nov-yinli",
    "Sheet4":       "dec-gaoming",
    "Sheet7":       "jan-tangspring",
    "Sheet9":       "feb-nansha",
    "Sheet5":       "mar-qingyuan",
    # apr-palm has no per-event sheet in the spreadsheet snapshot
}


def read_terry_event(wb, sheet_name):
    """Return list of {rank, name, points, stableford} from a per-event sheet.
    Sheets vary slightly: column headers are 'Rank | Name | Points | Score | Net Score | Stableford'
    or 'Rank | Name | Points | Score | Stableford'. We locate by header.
    """
    ws = wb[sheet_name]
    rows = []
    # Find the header row first
    header_row = None
    cols = {}
    for r in range(1, min(ws.max_row, 12)):
        for c in range(1, 9):
            v = ws.cell(row=r, column=c).value
            if v == "Rank":
                header_row = r
                # Capture all column positions on this row
                for cc in range(1, 9):
                    h = ws.cell(row=r, column=cc).value
                    if isinstance(h, str): cols[h.strip()] = cc
                break
        if header_row: break
    if not header_row: return []
    # Pre-scan: collect raw data rows. Then detect the Sheet4-style "stableford
    # in Points column" case — if the Stableford column is empty for nearly
    # every row but Points has values >13, treat Points as stableford for ALL
    # rows in this sheet (so the per-row swap is consistent).
    raw = []
    for r in range(header_row + 1, ws.max_row + 1):
        rank = ws.cell(row=r, column=cols.get("Rank", 2)).value
        nm   = ws.cell(row=r, column=cols.get("Name", 3)).value
        # Stop on rows that look like scramble/CTP labels in the Name col
        if isinstance(nm, str) and nm.strip() in (
                "Scramble Winners", "Drunkest Golfer", "Switch Hitter", "Shortest Drive",
                "CTP", "MH", "LD", "LC", "Rank"):
            break
        if not isinstance(nm, str) or not nm.strip():
            if raw: break  # blank row after data
            continue
        pts = ws.cell(row=r, column=cols.get("Points", 4)).value
        sb  = ws.cell(row=r, column=cols.get("Stableford", 7)).value
        raw.append({
            "rank_raw": rank,
            "name_raw": nm.strip(),
            "pts_raw": pts,
            "sb_raw":  sb,
        })

    # Sheet-level swap detection: if most rows have sb empty + pts > 13, swap.
    has_sb     = sum(1 for r in raw if isinstance(r["sb_raw"], (int, float)))
    has_big_p  = sum(1 for r in raw if isinstance(r["pts_raw"], (int, float)) and r["pts_raw"] > 13)
    sheet_swap = has_big_p > has_sb  # e.g. Sheet4: 17 big pts, ~1 sb

    # Compute rank from row position when Terry left the rank column blank
    auto_rank = sum(1 for r in raw if not isinstance(r["rank_raw"], int)) > 0

    for i, r in enumerate(raw):
        rank = r["rank_raw"] if isinstance(r["rank_raw"], int) else (i + 1)
        if sheet_swap:
            sb_val  = r["pts_raw"] if isinstance(r["pts_raw"], (int, float)) else None
            pts_val = None  # Terry didn't fill rank points on swapped sheets
        else:
            sb_val  = r["sb_raw"]
            pts_val = r["pts_raw"]
        rows.append({
            "rank": rank,
            "name_raw": r["name_raw"],
            "name": "",
            "seasonPts": int(pts_val) if isinstance(pts_val, (int, float)) else None,
            "stableford": int(sb_val) if isinstance(sb_val, (int, float)) else None,
            "rankPtsMissing": pts_val is None,  # signals "Terry didn't fill these"
        })
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Read Terry's Standings sheet
# ─────────────────────────────────────────────────────────────────────────────

def read_terry_standings():
    wb = openpyxl.load_workbook(SHEET_PATH, data_only=True)
    ws = wb["Standings"]
    raw = []
    for row in ws.iter_rows(values_only=True):
        # Layout: . | rank | name | points | played | round_avg | dblpar | neagle | nbirdie
        if row[1] == "Rank": continue
        if not isinstance(row[1], int): continue
        raw.append({
            "rank": row[1],
            "name_raw": (row[2] or "").strip(),
            "points": int(row[3] or 0),
            "played": int(row[4] or 0),
            "roundAvg": float(row[5] or 0),
            "dblpar": int(row[6] or 0),
            "neagle": int(row[7] or 0),
            "nbirdie": int(row[8] or 0),
        })
    # Build name set so context-aware Matt disambiguation can fire correctly:
    # standings sheet has both "Matt D" AND a separate "Matt" row — that bare
    # "Matt" must mean Matthew SA, not Matt D.
    name_set = {r["name_raw"] for r in raw}
    rows = []
    for r in raw:
        r["name"] = normalize_terry(r["name_raw"], name_set)
        rows.append(r)
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Diff & report
# ─────────────────────────────────────────────────────────────────────────────

def fmt_diff(label, app_val, terry_val):
    if app_val == terry_val: return None
    return f"  {label}: app={app_val} terry={terry_val} Δ={app_val - terry_val:+d}" if isinstance(app_val,int) else f"  {label}: app={app_val} terry={terry_val}"


def per_event_app_results(season):
    """Returns {event_id: [ {rank, name, gross, stableford, seasonPts, dblpar, neagle, nbirdie, cb} sorted ]}."""
    rank_pts = season.get("rankPoints") or DEFAULT_RANK_POINTS
    out = {}
    for ev in season["events"]:
        scored = []
        for p in ev["players"]:
            nm = canon_name(p["name"])
            sb = gross_total = dblpar = neagle = nbirdie = 0
            cb = []
            cb_tot = 0
            for i, hole in enumerate(ev["holes"]):
                g = p["gross"][i] if i < len(p["gross"]) else None
                if g is None or g == "": continue
                shots = shots_received(p["hcp"], hole["si"])
                pts = stableford_points(g, hole["par"], shots)
                if pts is not None: sb += pts
                gross_total += g
                diff = (g - shots) - hole["par"]
                if g >= 2 * hole["par"]: dblpar += 1
                if diff <= -2: neagle += 1
                if diff == -1: nbirdie += 1
            for i in range(len(ev["holes"]) - 1, -1, -1):
                g = p["gross"][i] if i < len(p["gross"]) else None
                cb_tot += g if isinstance(g, (int, float)) else 0
                cb.append(cb_tot)
            scored.append({"name": nm, "stableford": sb, "gross": gross_total,
                           "dblpar": dblpar, "neagle": neagle, "nbirdie": nbirdie, "cb": cb,
                           "hcp": p["hcp"]})
        scored.sort(key=lambda r: (-r["stableford"], *r["cb"]))
        for rank, r in enumerate(scored):
            r["rank"] = rank + 1
            r["seasonPts"] = rank_pts[rank] if rank < len(rank_pts) else 0
        out[ev["id"]] = scored
    return out


def cmd_per_event(season):
    print("\n=========== PER-EVENT DRILLDOWN ============")
    app_per_event = per_event_app_results(season)
    wb = openpyxl.load_workbook(SHEET_PATH, data_only=True)

    issues = []
    for ev in season["events"]:
        eid = ev["id"]
        # Find the matching Terry sheet (if any)
        terry_sheet = next((s for s, e in SHEET_TO_EVENT.items() if e == eid), None)
        print(f"\n── {eid} ({ev['date']}, {ev['courseName']}) ──")
        if terry_sheet is None:
            print(f"   (no Terry per-event sheet — skipping diff)")
            continue
        terry_rows = read_terry_event(wb, terry_sheet)
        # Normalize Terry names against this event's roster
        ev_roster = {canon_name(p["name"]) for p in ev["players"]}
        for tr in terry_rows:
            tr["name"] = normalize_terry(tr["name_raw"], ev_roster)
        terry_by_name = {r["name"]: r for r in terry_rows}
        app_by_name   = {r["name"]: r for r in app_per_event[eid]}

        # Tabular comparison
        print(f"  Source sheet: {terry_sheet}")
        print(f"  {'Rank':>4} {'Player':<14} {'SF(app/terry)':<16} {'SeasonPts(app/terry)':<22} {'gross':<6} {'hcp':<4}  Notes")
        for r in app_per_event[eid]:
            t = terry_by_name.get(r["name"])
            if t is None:
                notes = "TERRY-MISSING"
                t_sb = t_pts = "-"
            else:
                notes = []
                if r["stableford"] != t["stableford"] and t["stableford"] is not None:
                    notes.append(f"sf-Δ{r['stableford']-t['stableford']:+d}")
                if not t.get("rankPtsMissing") and r["seasonPts"] != t["seasonPts"]:
                    notes.append(f"pts-Δ{r['seasonPts']-(t['seasonPts'] or 0):+d}")
                if t["rank"] is not None and r["rank"] != t["rank"]:
                    notes.append(f"rank app{r['rank']}/terry{t['rank']}")
                notes = " ".join(notes) or "ok"
                t_sb = t["stableford"] if t["stableford"] is not None else "-"
                t_pts = "-" if t.get("rankPtsMissing") else t["seasonPts"]
            print(f"  {r['rank']:>4} {r['name']:<14} {str(r['stableford'])+'/'+str(t_sb):<16} {str(r['seasonPts'])+'/'+str(t_pts):<22} {r['gross']:<6} {r['hcp']:<4}  {notes}")
            if t and (r["stableford"] != t["stableford"] or r["seasonPts"] != t["seasonPts"] or
                      (t["rank"] is not None and r["rank"] != t["rank"])):
                issues.append((eid, r["name"], r, t))

        # Players only in Terry's sheet
        only_t = [r for r in terry_rows if r["name"] not in app_by_name]
        for r in only_t:
            print(f"   --   {r['name']:<14} (only in Terry sheet) sf={r['stableford']} pts={r['seasonPts']}")

    print(f"\n=== PER-EVENT SUMMARY: {len(issues)} player-event diff{'s' if len(issues)!=1 else ''} ===")


def cmd_player(season, player_name):
    print(f"\n=========== PLAYER DRILLDOWN: {player_name} ============")
    app_per_event = per_event_app_results(season)
    total_sb = total_pts = total_played = 0
    print(f"  {'Event':<22} {'Rank':>4}  {'SF':>3}  {'Gross':>5}  {'SeasonPts':>9}  {'Bird':>4}  {'Eag':>3}  {'Blob':>4}")
    for ev in season["events"]:
        eid = ev["id"]
        rec = next((r for r in app_per_event[eid] if r["name"] == player_name), None)
        if not rec:
            print(f"  {eid:<22} did not play")
            continue
        total_sb += rec["stableford"]; total_pts += rec["seasonPts"]; total_played += 1
        print(f"  {eid:<22} {rec['rank']:>4}  {rec['stableford']:>3}  {rec['gross']:>5}  {rec['seasonPts']:>9}  {rec['nbirdie']:>4}  {rec['neagle']:>3}  {rec['dblpar']:>4}")
    # Best 7
    rows = aggregate_season(season)
    target = next((r for r in rows if r["name"] == player_name), None)
    if target:
        print(f"\n  TOTAL  played={target['played']}  best7={target['best7']}  perfect10Bonus={target['perfect10Bonus']}  pts={target['points']}  seedPoints={target['seedPoints']}")


def main():
    if not SEASON_JSON.exists():
        print(f"Missing {SEASON_JSON}", file=sys.stderr); sys.exit(1)
    if not SHEET_PATH.exists():
        print(f"Missing {SHEET_PATH}", file=sys.stderr); sys.exit(1)

    season = json.loads(SEASON_JSON.read_text(encoding="utf-8"))
    excluded = set()
    do_per_event = False
    player_focus = None
    for arg in sys.argv[1:]:
        if arg.startswith("--exclude="):
            excluded = {x.strip() for x in arg.split("=", 1)[1].split(",") if x.strip()}
        elif arg in ("--per-event", "-e"):
            do_per_event = True
        elif arg.startswith("--player="):
            player_focus = arg.split("=", 1)[1]
    if excluded:
        before = len(season["events"])
        season["events"] = [e for e in season["events"] if e["id"] not in excluded]
        print(f"Excluding events {sorted(excluded)} — {before} -> {len(season['events'])} events")
    if do_per_event:
        cmd_per_event(season)
        if not player_focus: return
    if player_focus:
        cmd_player(season, player_focus)
        return
    print(f"Loaded season: {season['name']} — {len(season['events'])} events")
    for ev in season["events"]:
        print(f"  {ev['id']:20s} {ev['date']}  {ev['courseName']:25s}  players={len(ev['players'])}")

    print("\n— Computing standings via app logic —")
    app_rows = aggregate_season(season)

    print("— Reading Terry's Standings sheet —")
    terry_rows = read_terry_standings()
    terry_by_name = {r["name"]: r for r in terry_rows}
    app_by_name   = {r["name"]: r for r in app_rows}

    # Compare overlap
    print("\n=== STANDINGS DIFF ===")
    print(f"{'Rank':>4} {'Player':<14} {'Pts(app/terry)':<18} {'Play':<10} {'DblP':<10} {'NEag':<10} {'NBird':<10}")
    print("-" * 90)
    issues = 0
    for app_rank, app_r in enumerate(app_rows, 1):
        nm = app_r["name"]
        t = terry_by_name.get(nm)
        if t is None:
            print(f"{app_rank:>4} {nm:<14} APP-ONLY pts={app_r['points']} played={app_r['played']}")
            issues += 1
            continue
        diffs = []
        if app_r["points"]   != t["points"]:   diffs.append(f"pts={app_r['points']}/{t['points']}(Δ{app_r['points']-t['points']:+d})")
        if app_r["played"]   != t["played"]:   diffs.append(f"played={app_r['played']}/{t['played']}")
        if app_r["dblpar"]   != t["dblpar"]:   diffs.append(f"dblp={app_r['dblpar']}/{t['dblpar']}(Δ{app_r['dblpar']-t['dblpar']:+d})")
        if app_r["neagle"]   != t["neagle"]:   diffs.append(f"neag={app_r['neagle']}/{t['neagle']}(Δ{app_r['neagle']-t['neagle']:+d})")
        if app_r["nbirdie"]  != t["nbirdie"]:  diffs.append(f"nbird={app_r['nbirdie']}/{t['nbirdie']}(Δ{app_r['nbirdie']-t['nbirdie']:+d})")
        # Round average — float, allow tiny tolerance
        if abs(app_r["roundAvg"] - t["roundAvg"]) > 0.005:
            diffs.append(f"avg={app_r['roundAvg']:.3f}/{t['roundAvg']:.3f}")
        if app_rank != t["rank"]:
            diffs.append(f"rank=app{app_rank}/terry{t['rank']}")
        if diffs:
            issues += 1
            print(f"{app_rank:>4} {nm:<14} [DIFF] {' | '.join(diffs)}")
        else:
            print(f"{app_rank:>4} {nm:<14} OK   pts={app_r['points']} played={app_r['played']}")

    # Players in Terry's sheet but not in app
    only_terry = [r for r in terry_rows if r["name"] not in app_by_name]
    if only_terry:
        print(f"\n— TERRY-ONLY (not in season-4.json): {len(only_terry)} —")
        for r in only_terry:
            print(f"  rank={r['rank']:>2} {r['name']:<14} pts={r['points']} played={r['played']}")

    print(f"\n=== SUMMARY: {issues} discrepanc{'y' if issues==1 else 'ies'} flagged ===")


if __name__ == "__main__":
    main()
