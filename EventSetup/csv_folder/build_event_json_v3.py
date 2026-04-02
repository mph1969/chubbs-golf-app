#!/usr/bin/env python3
import csv
import json
import sys
from pathlib import Path
from openpyxl import load_workbook

def load_csv(path):
    with open(path, 'r', encoding='utf-8-sig', newline='') as f:
        return list(csv.DictReader(f))

def load_xlsx(path):
    wb = load_workbook(path, data_only=True)
    out = {}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(x).strip() if x is not None else '' for x in rows[0]]
        data = []
        for row in rows[1:]:
            if not any(v not in (None, '') for v in row):
                continue
            rec = {}
            for h, v in zip(headers, row):
                if h:
                    rec[h] = '' if v is None else v
            data.append(rec)
        out[ws.title] = data
    return out

def first_nonempty(*values):
    for value in values:
        text = str(value or '').strip()
        if text:
            return text
    return ''

def build_from_tables(tables):
    event = (tables.get('Event') or [{}])[0]

    players = []
    for row in tables.get('Players', []):
        if not str(row.get('playerId','')).strip() or not any(str(row.get(k,'')).strip() for k in ['displayName','fullName','shortName']):
            continue
        players.append({
            'playerId': str(row.get('playerId','')).strip(),
            'displayName': str(row.get('displayName','')).strip(),
            'fullName': str(row.get('fullName','')).strip(),
            'shortName': str(row.get('shortName','')).strip(),
            'playingHandicap': int(float(row.get('playingHandicap',0) or 0)),
            'defaultTee': str(row.get('defaultTee','')).strip(),
            'wechatName': str(row.get('wechatName','')).strip(),
        })

    day1 = []
    for row in tables.get('Day1 Scramble', []):
        if not str(row.get('teamId','')).strip() or not str(row.get('teamName','')).strip():
            continue
        pids = [str(row.get(f'player{i}Id','')).strip() for i in range(1,5)]
        player_ids = [p for p in pids if p]
        scorer = first_nonempty(row.get('scorerPlayerId',''), player_ids[0] if player_ids else '')
        day1.append({
            'teamId': str(row.get('teamId','')).strip(),
            'teamName': str(row.get('teamName','')).strip(),
            'scorerPlayerId': scorer,
            'playerIds': player_ids,
        })

    day2 = []
    for row in tables.get('Day2 Groups', []):
        if not str(row.get('groupId','')).strip() or not str(row.get('groupName','')).strip():
            continue
        pids = [str(row.get(f'player{i}Id','')).strip() for i in range(1,5)]
        player_ids = [p for p in pids if p]
        scorer = first_nonempty(row.get('scorerPlayerId',''), player_ids[0] if player_ids else '')
        day2.append({
            'groupId': str(row.get('groupId','')).strip(),
            'groupName': str(row.get('groupName','')).strip(),
            'teeTime': str(row.get('teeTime','')).strip(),
            'scorerPlayerId': scorer,
            'playerIds': player_ids,
        })

    return {
        'version': 2,
        'event': {
            'eventId': str(event.get('eventId','')).strip(),
            'name': str(event.get('name','')).strip(),
            'displayName': str(event.get('displayName','')).strip(),
            'satDate': str(event.get('satDate','')).strip(),
            'satTime': str(event.get('satTime','')).strip(),
            'sunDate': str(event.get('sunDate','')).strip(),
            'sunTime': str(event.get('sunTime','')).strip(),
            'day1CourseId': first_nonempty(event.get('day1CourseId',''), event.get('courseId','')),
            'day1CourseLabel': first_nonempty(event.get('day1CourseLabel',''), event.get('courseLabel','')),
            'day1PalmIslandFront': first_nonempty(event.get('day1PalmIslandFront',''), event.get('palmIslandFront','')),
            'day1PalmIslandBack': first_nonempty(event.get('day1PalmIslandBack',''), event.get('palmIslandBack','')),
            'day2CourseId': first_nonempty(event.get('day2CourseId',''), event.get('courseId','')),
            'day2CourseLabel': first_nonempty(event.get('day2CourseLabel',''), event.get('courseLabel','')),
            'day2PalmIslandFront': first_nonempty(event.get('day2PalmIslandFront',''), event.get('palmIslandFront','')),
            'day2PalmIslandBack': first_nonempty(event.get('day2PalmIslandBack',''), event.get('palmIslandBack','')),
            'currency': str(event.get('currency','RMB')).strip() or 'RMB',
            'notes': str(event.get('notes','')).strip(),
        },
        'options': {
            'threePuttPoker': True,
            'pokerMode': 'digital',
            'ante': 50,
            'penalty3': 10,
            'penalty4plus': 20,
        },
        'players': players,
        'day1ScrambleTeams': day1,
        'day2Groups': day2,
    }

def build_from_csv_folder(folder):
    folder = Path(folder)
    return build_from_tables({
        'Event': load_csv(folder / 'cpi_event.csv'),
        'Players': load_csv(folder / 'cpi_players.csv'),
        'Day1 Scramble': load_csv(folder / 'cpi_day1_scramble_teams.csv'),
        'Day2 Groups': load_csv(folder / 'cpi_day2_groups.csv'),
    })

def main():
    if len(sys.argv) < 3:
        print('Usage: python build_event_json_v3.py template.xlsx output.json')
        print('   or: python build_event_json_v3.py folder_with_csvs output.json')
        raise SystemExit(1)
    source = Path(sys.argv[1])
    output = Path(sys.argv[2])
    if source.is_dir():
        data = build_from_csv_folder(source)
    elif source.suffix.lower() == '.xlsx':
        data = build_from_tables(load_xlsx(source))
    else:
        raise SystemExit('Source must be an .xlsx file or a folder containing the CSV templates.')
    output.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding='utf-8')
    print(f'Wrote {output}')

if __name__ == '__main__':
    main()
